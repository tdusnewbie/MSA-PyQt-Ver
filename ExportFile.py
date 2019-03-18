import sys

from openpyxl import *
from openpyxl.styles import Color, PatternFill, Font, DEFAULT_FONT
from openpyxl.cell import Cell
from PyQt5 import QtWidgets, QtGui, QtCore
from Model import Database

BLUE = PatternFill(start_color='9898e6', end_color='9898e6',fill_type='solid')
YELLOW = PatternFill(start_color='FFFF00', end_color='FFFF00',fill_type='solid')
titleFont = Font(name = "Calibri",sz = 14,family=2,b=True, i = True)


def createTitleTypeItem(sheet, item):
    sheet.cell(1,1,"Tên mặt hàng").fill = YELLOW
    sheet.cell(1,1).font = titleFont

    sheet.merge_cells(start_row = 1,end_row = 1, start_column = 1, end_column = 2)
    sheet.cell(1,3,item.name)
    sheet.merge_cells(start_row = 1,end_row = 1, start_column = 3, end_column = 4)

    sheet.cell(2,1,"ID").fill = YELLOW
    sheet.cell(2,1).font = titleFont

    sheet.merge_cells(start_row = 2,end_row = 2, start_column = 1, end_column = 2)
    sheet.cell(2,3,item.idItem)
    sheet.merge_cells(start_row = 2,end_row = 2, start_column = 3, end_column = 4)

    sheet.cell(3,1,"Tên loại hàng").fill = BLUE
    sheet.cell(3,1).font = titleFont
    
    sheet.cell(3,2,"ID loại hàng").fill = BLUE
    sheet.cell(3,2).font = titleFont

    sheet.cell(3,3,"Số lượng").fill = BLUE
    sheet.cell(3,3).font = titleFont

    sheet.cell(3,4,"Giá lẻ").fill = BLUE
    sheet.cell(3,4).font = titleFont

    sheet.cell(3,5,"Giá sỉ").fill = BLUE
    sheet.cell(3,5).font = titleFont

    sheet.cell(3,6,"Giá vốn").fill = BLUE
    sheet.cell(3,6).font = titleFont

    sheet.cell(3,7,"Ghi chú").fill = BLUE
    sheet.cell(3,7).font = titleFont

    sheet.cell(3,8,"Hình ảnh").fill = BLUE
    sheet.cell(3,8).font = titleFont

def createdTitleBill(sheet, createdDate, numberBill):
    sheet.cell(1,1,"Ngày tạo").fill = YELLOW
    sheet.cell(1,1).font = titleFont

    sheet.merge_cells(start_row = 1,end_row = 1, start_column = 1, end_column = 2)
    sheet.cell(1,3,createdDate)
    sheet.merge_cells(start_row = 1,end_row = 1, start_column = 3, end_column = 4)

    sheet.cell(2,1,"Số lượng").fill = YELLOW
    sheet.cell(2,1).font = titleFont

    sheet.merge_cells(start_row = 2,end_row = 2, start_column = 1, end_column = 2)
    sheet.cell(2,3,numberBill)
    sheet.merge_cells(start_row = 2,end_row = 2, start_column = 3, end_column = 4)

    sheet.cell(3,1,"Giờ tạo").fill = BLUE
    sheet.cell(3,1).font = titleFont
    
    sheet.cell(3,2,"Tên mặt hàng").fill = BLUE
    sheet.cell(3,2).font = titleFont

    sheet.cell(3,3,"Tên loại hàng").fill = BLUE
    sheet.cell(3,3).font = titleFont

    sheet.cell(3,4,"ID").fill = BLUE
    sheet.cell(3,4).font = titleFont
    
    sheet.cell(3,5,"Số lượng").fill = BLUE
    sheet.cell(3,5).font = titleFont
    
    sheet.cell(3,6,"Phương thức bán").fill = BLUE
    sheet.cell(3,6).font = titleFont

    sheet.cell(3,7,"Thành giá").fill = BLUE
    sheet.cell(3,7).font = titleFont

    sheet.cell(3,8,"Người tạo").fill = BLUE
    sheet.cell(3,8).font = titleFont

    sheet.cell(3,9,"Ghi chú").fill = BLUE
    sheet.cell(3,9).font = titleFont

def adjustColumn(sheet):
    for col in sheet.columns:
        max_length = 0
        column = col[0].column # Get the column name
        for cell in col:
            try: # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        sheet.column_dimensions[column].width = adjusted_width

def addTypeIntoSheet(sheet,listType):
    numType = len(listType)
    for row in range(4,numType+4):
        typeItem = listType[row-1-3]
        for col in range(1,9):
            if col == 1:
                sheet.cell(row,col,typeItem.name)
            if col == 2:
                sheet.cell(row,col,typeItem.idType)
            if col == 3:
                sheet.cell(row,col,typeItem.amount)
            if col == 4:
                sheet.cell(row,col,typeItem.unitPrice)
            if col == 5:
                sheet.cell(row,col,typeItem.wholePrice)
            if col == 6:
                sheet.cell(row,col,typeItem.originPrice)
            if col == 7:
                sheet.cell(row,col,typeItem.notice)
            if col == 8:
                sheet.cell(row,col,typeItem.imagePath)

def addBillIntoSheet(sheet,listBill):
    numBill = len(listBill)
    for row in range(4,numBill+4):
        bill = listBill[row-1-3]
        for col in range(1,9):
            if col == 1:
                sheet.cell(row,col,bill.createdTime)
            if col == 2:
                sheet.cell(row,col,bill.nameItem)
            if col == 3:
                sheet.cell(row,col,bill.nameType)
            if col == 4:
                sheet.cell(row,col,bill.idBill)
            if col == 5:
                sheet.cell(row,col,bill.amount)
            if col == 6:
                sheet.cell(row,col,bill.saleType)
            if col == 7:
                sheet.cell(row,col,bill.price)
            if col == 8:
                sheet.cell(row,col,bill.createdUser)
            if col == 9:
                sheet.cell(row,col,bill.notice)

def createSheetAsItem(wb,item):
    ws = wb.create_sheet(item.name)
    return ws

def createSheetAsCreatedDate(wb,createdDate):
    ws = wb.create_sheet(createdDate)
    return ws

def exportFileItem(pathFile):
    if len(pathFile) == 0:
        return
    wb = Workbook()
    ws = wb.active

    data = Database()
    listItem = data.getListItem()

    for item in listItem:
        ws = createSheetAsItem(wb,item)
        createTitleTypeItem(ws, item)
        addTypeIntoSheet(ws,item.listType)
        adjustColumn(ws)

    wb.remove_sheet(wb.active)
    wb.save(pathFile)

def exportFileBill(pathFile):
    # if pathFile == None:
    #     return
    # wb = Workbook()
    # ws = wb.active

    # data = Database()
    # listBill = data.getAllListBill()

    # listDate = []

    # for bill in listBill:
    #     ws = createSheetAsItem(wb,item)
    #     createTitleTypeItem(ws, item)
    #     addTypeIntoSheet(ws,item.listType)
    #     adjustColumn(ws)

    # wb.remove_sheet(wb.active)
    # wb.save(pathFile)
    pass

if __name__ == "__main__":
    app = QtWidgets.QApplication(sys.argv)
    pathFile,_ = QtWidgets.QFileDialog.getSaveFileName(None,"Save file",".",
            "Excel files (*.xlsx)")
    exportFileItem(pathFile)
    sys.exit(app.exec_())